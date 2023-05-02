from pathlib import Path
import openpyxl 
import os

# template of excel sheet
template = ["Reservation number" ,	"COMPANY NAME" , 	"NM" ,	"MS" ,	"Character" ,
"DATE" ,	"ROUTH" ,	"HK" ,	 "MS" , "ROUTH" ,  "DATE" ,	"APM" ,	"APE" ,]

# files
def file_is_exist(abs_path):
    my_file = Path(abs_path)
    if my_file.exists():
        return True
    else:
        return False

def create_file(abs_path):
    f = open(abs_path , 'w')
    f.write("2")
    f.close()
    

def get_current_directory(name_of_file):
    new_list = __file__.split('\\')
    new_list.pop()
    new_list.append(name_of_file)
    abs_path = "\\".join(new_list)
    return abs_path
    

def get_index_from_file(name_of_file):
    index = 2
    abs_path = get_current_directory(name_of_file)
    if file_is_exist(abs_path):
        f = open(abs_path , 'r')
        index = f.read()
        f.close()
    else:
        create_file(abs_path)
    return index  

def increment_index_of_file(name_of_file , index):
    abs_path = get_current_directory(name_of_file)
    f = open(abs_path , 'w')
    f.write(str(index + 1))
    f.close()

# sheet
def put_data_into_row_in_excel(sheet , index , data ):
    for col , value in enumerate(data):
        sheet.cell(row = index , column = col + 1 , value = value)
        

def create_template_of_excel(sheet, template):
    for i in range(0 , len(template)):
        sheet.cell(row = 1 , column = i + 1 , value = template[i])
        sheet.cell(row = 1, column =  i + 1).font = openpyxl.styles.Font(size = 14, bold = True)
    # Set the default row height to 20 and the default column width to 10
    sheet.sheet_format.defaultRowHeight = 25
    sheet.sheet_format.defaultColWidth = 30    

def get_excel_sheet(name_of_file):
    sheet = None
    abs_path = get_current_directory(name_of_file)
    # Check if file exists
    if file_is_exist(abs_path):
        # Create new workbook object
        wb = openpyxl.load_workbook(abs_path)
        # active sheet 
        sheet = wb.active
        
    else:
        # Create new workbook object
        wb = openpyxl.Workbook()
        # active sheet 
        sheet = wb.active
        # Set default row height and column width
        
        # create template
        create_template_of_excel(sheet , template)
        # Save workbook to file
        wb.save(abs_path)
    
    return wb , sheet
    

        
def fill_data_into_excel_sheet(data):
    name_of_file = "number_of_rows.txt"
    index = int(get_index_from_file(name_of_file))
    
    name_of_excel_sheet = 'reservations.xlsx'
    
    wb , sheet = get_excel_sheet(name_of_excel_sheet)
    
    put_data_into_row_in_excel(sheet , index , data)
    wb.save(get_current_directory(name_of_excel_sheet))
    
    increment_index_of_file(name_of_file , index)
    
    






